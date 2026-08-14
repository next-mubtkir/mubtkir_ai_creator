"""أداة الاستيراد: تعالج الملف كاملًا بالكود، والنموذج يبني خريطة الحقول فقط.

التصميم يحلّ مشكلة التكلفة والحجم معًا:
- النموذج يرى رؤوس الأعمدة وعيّنة صغيرة فقط (استدعاء واحد)
- الكود يعالج آلاف الصفوف بلا أي استدعاء إضافي للنموذج
- التحقق من الحقول الإجبارية وحقول الربط يتم لكل صف قبل الكتابة
"""

import csv
import datetime
import io
import json
import os
import time

import frappe
from frappe.utils import now_datetime

from mubtkir_ai_creator.lib import llm, tools
from mubtkir_ai_creator.lib.agent import _dump, log_action
from mubtkir_ai_creator.lib.client import FrappeSiteClient

SAMPLE_ROWS = 5          # ما يراه النموذج فقط
MAX_TOTAL_ROWS = 20000   # حد أمان لحجم الملف
COMMIT_EVERY = 25        # حفظ التقدّم دوريًا أثناء التنفيذ الطويل

# صيغ التاريخ الشائعة بالملفات — تُجرَّب بالترتيب حتى تنجح واحدة
DATE_FORMATS = (
    "%Y-%m-%d", "%Y/%m/%d",
    "%d-%m-%Y", "%d/%m/%Y",
    "%m-%d-%Y", "%m/%d/%Y",
    "%d.%m.%Y", "%Y-%m-%d %H:%M:%S",
)


def _try_parse_date(value):
    """يحاول عدة صيغ تاريخ شائعة ويرجع yyyy-mm-dd الموحّدة، أو None لو فشلت كلها."""
    if value in (None, ""):
        return None
    if isinstance(value, (datetime.date, datetime.datetime)):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s:
        return None
    for fmt in DATE_FORMATS:
        try:
            return datetime.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


# ---------------- قراءة الملف ----------------

def read_rows(file_url):
    """قراءة كل صفوف الملف كقواميس {اسم العمود: القيمة}."""
    name = frappe.db.get_value("File", {"file_url": file_url}, "name")
    if not name:
        frappe.throw(f"لم يُعثر على الملف: {file_url}")

    doc = frappe.get_doc("File", name)
    ext = os.path.splitext(doc.file_name or "")[1].lower()
    content = doc.get_content()
    raw = content.encode("utf-8") if isinstance(content, str) else content

    if ext in (".xlsx", ".xlsm"):
        headers, rows = _read_xlsx_rows(raw)
    elif ext == ".csv":
        headers, rows = _read_csv_rows(raw)
    else:
        frappe.throw(f"صيغة غير مدعومة للاستيراد: {ext}. المدعوم: xlsx و csv")

    if len(rows) > MAX_TOTAL_ROWS:
        frappe.throw(f"الملف يحتوي {len(rows)} صفًا — الحد الأقصى {MAX_TOTAL_ROWS}")

    return headers, rows


def read_rows_for_import(doc):
    """يقرأ صفوف الاستيراد من ملف مرفوع أو من رابط Google Sheet — أيهما محدد."""
    if doc.source_file:
        return read_rows(doc.source_file)
    if doc.google_sheet_url:
        return _read_google_sheet_rows(doc.google_sheet_url)
    frappe.throw("حدد ملف استيراد أو رابط Google Sheet")


def _read_google_sheet_rows(sheet_url):
    import re
    import requests

    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", sheet_url or "")
    if not m:
        frappe.throw("رابط Google Sheet غير صالح — لازم يكون بصيغة docs.google.com/spreadsheets/d/...")
    sheet_id = m.group(1)
    gid_match = re.search(r"[?#&]gid=(\d+)", sheet_url)
    gid = gid_match.group(1) if gid_match else "0"

    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
    try:
        resp = requests.get(export_url, timeout=30)
    except requests.RequestException as e:
        frappe.throw(f"تعذّر الوصول لملف Google Sheet: {e}")

    if resp.status_code != 200 or resp.text.lstrip().lower().startswith("<!doctype html"):
        frappe.throw(
            "تعذّر قراءة Google Sheet — تأكد أن صلاحية المشاركة \"يمكن لأي شخص لديه الرابط العرض\""
        )

    headers, rows = _read_csv_rows(resp.content)
    if len(rows) > MAX_TOTAL_ROWS:
        frappe.throw(f"الشيت يحتوي {len(rows)} صفًا — الحد الأقصى {MAX_TOTAL_ROWS}")
    return headers, rows


def _read_xlsx_rows(raw):
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    ws = wb.worksheets[0]  # الورقة الأولى فقط

    headers, rows = [], []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        values = ["" if c is None else str(c).strip() for c in row]
        if i == 0:
            headers = [h for h in values]
            continue
        if not any(v for v in values):
            continue
        rows.append({headers[j]: values[j] for j in range(min(len(headers), len(values))) if headers[j]})

    wb.close()
    return [h for h in headers if h], rows


def _read_csv_rows(raw):
    import csv

    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("cp1256", errors="replace")

    reader = csv.DictReader(io.StringIO(text))
    headers = [h.strip() for h in (reader.fieldnames or []) if h]
    rows = []
    for r in reader:
        clean = {(k or "").strip(): (str(v).strip() if v is not None else "") for k, v in r.items() if k}
        if any(clean.values()):
            rows.append(clean)
    return headers, rows


# ---------------- بناء الخريطة (الاستدعاء الوحيد للنموذج) ----------------

def analyze(name):
    doc = frappe.get_doc("AI Import", name)
    headers, rows = read_rows_for_import(doc)

    if not headers:
        frappe.throw("لم يُعثر على رؤوس أعمدة في الملف")

    client = FrappeSiteClient(doc.client_site)
    spec = tools.describe_required(client, doc.target_doctype, with_options=True)
    meta = client.get_meta(doc.target_doctype).get("data", {}) or {}

    fields_brief = [
        {
            "fieldname": f.get("fieldname"),
            "label": f.get("label"),
            "fieldtype": f.get("fieldtype"),
            "link_to": f.get("options") if f.get("fieldtype") == "Link" else None,
            "required": bool(f.get("reqd")),
        }
        for f in meta.get("fields", [])
        if f.get("fieldname") and f.get("fieldtype") not in ("Section Break", "Column Break", "Tab Break", "HTML", "Table")
    ][:120]

    sample = rows[:SAMPLE_ROWS]

    prompt = f"""اربط أعمدة ملف الاستيراد بحقول DocType «{doc.target_doctype}» في ERPNext.

أعمدة الملف: {json.dumps(headers, ensure_ascii=False)}

عيّنة من الصفوف: {json.dumps(sample, ensure_ascii=False)}

حقول الـ DocType المتاحة: {json.dumps(fields_brief, ensure_ascii=False)}

الحقول الإجبارية: {json.dumps([f['fieldname'] for f in spec['required_fields']], ensure_ascii=False)}

أعد JSON فقط بلا أي نص آخر وبلا علامات markdown، بهذا الشكل:
{{
  "mapping": {{"اسم العمود في الملف": "fieldname في ERPNext"}},
  "unmapped_columns": ["أعمدة لم تجد لها مقابلًا"],
  "missing_required": ["حقول إجبارية لا يوجد لها عمود في الملف"],
  "notes": "ملاحظات موجزة بالعربية"
}}"""

    result = llm.chat([{"role": "user", "content": prompt}], system="أنت محلل بيانات. أعد JSON صالحًا فقط.", heavy=True)

    text = (result.get("text") or "").strip()
    text = text.replace("```json", "").replace("```", "").strip()
    try:
        parsed = json.loads(text)
    except ValueError:
        start, end = text.find("{"), text.rfind("}")
        if start == -1 or end == -1:
            frappe.throw(f"تعذّر قراءة خريطة الحقول من النموذج. الرد: {text[:400]}")
        parsed = json.loads(text[start : end + 1])

    doc.db_set("detected_columns", json.dumps(headers, ensure_ascii=False))
    doc.db_set("sample_rows", json.dumps(sample, ensure_ascii=False, indent=2))
    doc.db_set("total_rows", len(rows))
    doc.db_set(
        "analysis_notes",
        f"أعمدة بلا مقابل: {'، '.join(parsed.get('unmapped_columns') or []) or 'لا يوجد'}\n"
        f"حقول إجبارية بلا عمود: {'، '.join(parsed.get('missing_required') or []) or 'لا يوجد'}\n"
        f"{parsed.get('notes') or ''}",
    )
    doc.db_set("status", "Mapping Ready")

    mapping = parsed.get("mapping") or {}
    sample_lookup = sample[0] if sample else {}
    doc.set("field_mapping", [])
    for column in headers:
        doc.append(
            "field_mapping",
            {
                "source_column": column,
                "target_fieldname": mapping.get(column, ""),
                "sample_value": str(sample_lookup.get(column, ""))[:140],
            },
        )
    doc.save(ignore_permissions=True)
    frappe.db.commit()

    return {"headers": headers, "total_rows": len(rows), "mapping": mapping, "notes": parsed.get("notes")}


# ---------------- تحويل الصفوف ----------------

def _date_fields_for(client, target_doctype):
    """أسماء الحقول من نوع Date/Datetime لدى الـ DocType الهدف — لتطبيع صيغ التاريخ المختلفة قبل الإرسال."""
    meta = client.get_meta(target_doctype).get("data", {}) or {}
    return {
        f.get("fieldname")
        for f in meta.get("fields", [])
        if f.get("fieldtype") in ("Date", "Datetime") and f.get("fieldname")
    }


def _build_docs(doc, rows, date_fields=None):
    mapping = {r.source_column: r.target_fieldname for r in (doc.field_mapping or []) if r.target_fieldname}
    statics = json.loads(doc.static_values or "{}")
    date_fields = date_fields or set()

    out = []
    for i, row in enumerate(rows, start=2):  # الصف 1 رؤوس الأعمدة
        payload = dict(statics)
        date_issues = []
        for column, fieldname in mapping.items():
            if not fieldname:
                continue
            value = row.get(column)
            if value in (None, ""):
                continue
            if fieldname in date_fields:
                parsed_date = _try_parse_date(value)
                if parsed_date:
                    payload[fieldname] = parsed_date
                else:
                    date_issues.append(f"تعذّر فهم التاريخ في {fieldname}: «{value}»")
                    payload[fieldname] = value
            else:
                payload[fieldname] = value
        out.append({"row_number": i, "data": payload, "date_issues": date_issues})
    return out


# ---------------- المعاينة ----------------

def preview(name):
    doc = frappe.get_doc("AI Import", name)
    _, rows = read_rows_for_import(doc)
    client = FrappeSiteClient(doc.client_site)
    date_fields = _date_fields_for(client, doc.target_doctype)
    prepared = _build_docs(doc, rows, date_fields)

    # الحقول الإجبارية تُفحص من التعريف مرة واحدة، وحقول الربط تُفحص بقيم فريدة فقط
    spec = tools.describe_required(client, doc.target_doctype, with_options=True)
    required = [f for f in spec["required_fields"] if not f.get("default")]
    link_fields = {f["fieldname"]: f["link_to"] for f in spec["required_fields"] if f.get("link_to")}

    meta = client.get_meta(doc.target_doctype).get("data", {}) or {}
    for f in meta.get("fields", []):
        if f.get("fieldtype") == "Link" and f.get("options"):
            link_fields[f.get("fieldname")] = f.get("options")

    # جمع القيم الفريدة لكل حقل ربط والتحقق منها دفعة واحدة (بدل صف صف)
    unique_values = {}
    for item in prepared:
        for fieldname, target in link_fields.items():
            v = item["data"].get(fieldname)
            if v:
                unique_values.setdefault(fieldname, set()).add(v)

    invalid_links = {}
    for fieldname, values in unique_values.items():
        target = link_fields[fieldname]
        for v in values:
            try:
                found = client.get_list(target, fields=["name"], filters={"name": v}, limit=1).get("data") or []
            except Exception:
                found = []
            if not found:
                invalid_links.setdefault(fieldname, {"doctype": target, "values": []})
                invalid_links[fieldname]["values"].append(v)

    # فحص الصفوف — بما فيها صيغة التاريخ (تُفحص هنا بدل انتظار فشل التنفيذ الفعلي)
    issues, ok_count = [], 0
    for item in prepared:
        row_issues = list(item.get("date_issues") or [])
        for f in required:
            if not item["data"].get(f["fieldname"]):
                row_issues.append(f"حقل إجباري ناقص: {f.get('label') or f['fieldname']}")
        for fieldname, info in invalid_links.items():
            v = item["data"].get(fieldname)
            if v and v in info["values"]:
                row_issues.append(f"قيمة ربط غير موجودة في {fieldname}: {v}")

        if row_issues:
            issues.append({"row": item["row_number"], "issues": row_issues, "sample": item["data"]})
        else:
            ok_count += 1

    for fieldname, info in invalid_links.items():
        try:
            avail = client.get_list(info["doctype"], fields=["name"], limit=15).get("data") or []
            info["available_options"] = [a.get("name") for a in avail]
        except Exception:
            info["available_options"] = []

    summary = f"صفوف جاهزة: {ok_count} | صفوف بها مشاكل: {len(issues)} | الإجمالي: {len(prepared)}"

    doc.db_set("total_rows", len(prepared))
    doc.db_set("valid_rows", ok_count)
    doc.db_set("invalid_rows", len(issues))
    doc.db_set(
        "preview_result",
        json.dumps(
            {"summary": summary, "invalid_links": invalid_links, "row_issues": issues[:200]},
            ensure_ascii=False,
            indent=2,
        ),
    )
    doc.db_set("status", "Pending Approval")
    frappe.db.commit()

    return {"summary": summary, "valid": ok_count, "invalid": len(issues), "invalid_links": invalid_links, "issues": issues[:50]}


@frappe.whitelist()
def download_failure_rows(name):
    """يبني ملف CSV يحتوي فقط الصفوف الفاشلة (من آخر معاينة أو تنفيذ) وسبب كل فشل، ويرجع رابط تنزيله."""
    frappe.only_for(["System Manager", "AI Creator User", "AI Creator Supervisor"])
    doc = frappe.get_doc("AI Import", name)

    rows = []
    try:
        parsed = json.loads(doc.preview_result or "{}")
        rows = parsed.get("row_issues") or []
    except ValueError:
        rows = []

    if not rows and doc.failure_report:
        # تنفيذ فعلي (execute) يخزن failure_report كنص — نحوّله لصفوف بسيطة
        for line in (doc.failure_report or "").splitlines():
            if line.strip():
                rows.append({"row": "", "issues": [line.strip()], "sample": {}})

    if not rows:
        frappe.throw("لا توجد صفوف فاشلة مسجّلة لهذا الاستيراد")

    buf = io.StringIO()
    all_fields = sorted({k for r in rows for k in (r.get("sample") or {}).keys()})
    writer = csv.writer(buf)
    writer.writerow(["row_number", "issues", *all_fields])
    for r in rows:
        sample = r.get("sample") or {}
        writer.writerow([r.get("row", ""), " | ".join(r.get("issues") or []), *[sample.get(f, "") for f in all_fields]])

    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": f"{doc.name}-failed-rows.csv",
        "content": buf.getvalue(),
        "attached_to_doctype": "AI Import",
        "attached_to_name": doc.name,
        "is_private": 1,
    })
    file_doc.insert(ignore_permissions=True)
    frappe.db.commit()
    return {"file_url": file_doc.file_url}


# ---------------- التنفيذ ----------------

def enqueue_execute(name):
    """التنفيذ في الخلفية — ملف كبير يتجاوز مهلة طلب HTTP."""
    doc = frappe.get_doc("AI Import", name)
    if doc.status != "Approved":
        frappe.throw("لا يمكن التنفيذ قبل الاعتماد")

    doc.db_set("status", "Queued")
    doc.db_set("processed_rows", 0)
    frappe.db.commit()

    frappe.enqueue(
        "mubtkir_ai_creator.lib.importer.execute",
        queue="long",
        timeout=7200,
        name=name,
        enqueue_after_commit=True,
    )
    return {"status": "Queued"}


def execute(name):
    started = time.time()
    doc = frappe.get_doc("AI Import", name)
    _, rows = read_rows_for_import(doc)
    client = FrappeSiteClient(doc.client_site)
    date_fields = _date_fields_for(client, doc.target_doctype)
    prepared = _build_docs(doc, rows, date_fields)

    skip_invalid = bool(doc.skip_invalid_rows)

    doc.db_set("status", "Executing")
    frappe.db.commit()

    success = failed = 0
    failures = []

    for idx, item in enumerate(prepared, start=1):
        try:
            out = client.create_doc(doc.target_doctype, item["data"])
            created = (out or {}).get("data", {}).get("name")
            success += 1
        except Exception as e:
            failed += 1
            failures.append({"row": item["row_number"], "error": str(e)[:400]})
            if not skip_invalid:
                doc.db_set("status", "Failed")
                doc.db_set("error_message", f"توقف عند الصف {item['row_number']}: {str(e)[:500]}")
                break

        if idx % COMMIT_EVERY == 0:
            doc.db_set("processed_rows", idx)
            doc.db_set("success_count", success)
            doc.db_set("failed_count", failed)
            frappe.db.commit()

    doc.db_set("processed_rows", len(prepared))
    doc.db_set("success_count", success)
    doc.db_set("failed_count", failed)
    doc.db_set("failure_report", json.dumps(failures[:500], ensure_ascii=False, indent=2))

    if doc.status != "Failed":
        doc.db_set("status", "Completed" if failed == 0 else "Partially Failed")

    log_action(
        client_site=doc.client_site,
        site_url=client.site_url,
        tool_name="bulk_import",
        risk_level="high",
        tool_input=_dump({"import": doc.name, "doctype": doc.target_doctype, "rows": len(prepared)}),
        tool_output=_dump({"success": success, "failed": failed}),
        is_success=1 if failed == 0 else 0,
        duration_ms=int((time.time() - started) * 1000),
        error_message=None if failed == 0 else f"فشل {failed} صفًا — راجع تقرير الفشل",
    )

    frappe.db.commit()
    return {"status": doc.status, "success": success, "failed": failed}
