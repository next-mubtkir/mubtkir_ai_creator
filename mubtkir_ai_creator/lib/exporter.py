"""تصدير تخصيصات العميل (AI Templates) إلى Excel أو PDF."""

import io
import json

import frappe
from frappe.utils import now_datetime

from mubtkir_ai_creator.lib.client import FrappeSiteClient


def export_templates_excel(client_site=None, artifact_types=None, target_doctype=None):
    """تصدير القوالب المحفوظة في AI Template إلى ملف Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        frappe.throw("مكتبة openpyxl مطلوبة للتصدير — pip install openpyxl")

    filters = {}
    if client_site:
        filters["source_client"] = client_site
    if artifact_types:
        if isinstance(artifact_types, str):
            artifact_types = [artifact_types]
        filters["artifact_type"] = ["in", artifact_types]
    if target_doctype:
        filters["target_doctype"] = target_doctype

    templates = frappe.get_all(
        "AI Template",
        filters=filters,
        fields=["name", "title", "artifact_type", "source_client", "source_name",
                "target_doctype", "version", "deployable", "captured_on", "notes", "payload"],
        order_by="source_client asc, artifact_type asc, source_name asc",
        limit_page_length=0,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Templates"
    ws.sheet_view.rightToLeft = True

    headers = ["Template ID", "Title", "Type", "Source Client", "Source Name",
               "Target DocType", "Version", "Deployable", "Captured On", "Notes"]
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="3867AE", end_color="3867AE", fill_type="solid")

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")

    for i, t in enumerate(templates, 2):
        ws.cell(row=i, column=1, value=t.name)
        ws.cell(row=i, column=2, value=t.title)
        ws.cell(row=i, column=3, value=t.artifact_type)
        ws.cell(row=i, column=4, value=t.source_client)
        ws.cell(row=i, column=5, value=t.source_name)
        ws.cell(row=i, column=6, value=t.target_doctype)
        ws.cell(row=i, column=7, value=t.version)
        ws.cell(row=i, column=8, value="Yes" if t.deployable else "No")
        ws.cell(row=i, column=9, value=str(t.captured_on or ""))
        ws.cell(row=i, column=10, value=t.notes)

    # Sheet per artifact type with payload details
    by_type = {}
    for t in templates:
        by_type.setdefault(t.artifact_type, []).append(t)

    for atype, items in by_type.items():
        sheet_name = atype.replace("/", "-")[:30]
        ws2 = wb.create_sheet(title=sheet_name)
        ws2.sheet_view.rightToLeft = True
        ws2.cell(row=1, column=1, value="Source Name").font = Font(bold=True)
        ws2.cell(row=1, column=2, value="Client").font = Font(bold=True)
        ws2.cell(row=1, column=3, value="Payload (JSON)").font = Font(bold=True)

        for j, item in enumerate(items, 2):
            ws2.cell(row=j, column=1, value=item.source_name)
            ws2.cell(row=j, column=2, value=item.source_client)
            ws2.cell(row=j, column=3, value=item.payload or "")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"AI_Templates_{frappe.utils.now().replace(' ', '_').replace(':', '')}.xlsx"
    fpath = frappe.get_site_path("private", "files", fname)
    with open(fpath, "wb") as f:
        f.write(buf.read())

    fdoc = frappe.get_doc({
        "doctype": "File",
        "file_name": fname,
        "file_url": f"/private/files/{fname}",
        "is_private": 1,
    })
    fdoc.insert(ignore_permissions=True)
    frappe.db.commit()

    return {"file_url": fdoc.file_url, "file_name": fname, "template_count": len(templates)}


def export_templates_pdf(client_site=None, artifact_types=None, target_doctype=None):
    """تصدير القوالب إلى تقرير PDF منسّق."""
    filters = {}
    if client_site:
        filters["source_client"] = client_site
    if artifact_types:
        if isinstance(artifact_types, str):
            artifact_types = [artifact_types]
        filters["artifact_type"] = ["in", artifact_types]
    if target_doctype:
        filters["target_doctype"] = target_doctype

    templates = frappe.get_all(
        "AI Template",
        filters=filters,
        fields=["name", "title", "artifact_type", "source_client", "source_name",
                "target_doctype", "version", "captured_on", "deployable", "notes", "payload"],
        order_by="source_client asc, artifact_type asc",
        limit_page_length=0,
    )

    html_parts = [
        f"""<html dir="rtl"><head><meta charset="utf-8">
        <style>
            body{{font-family:Arial,sans-serif;font-size:12px;color:#243B63;padding:20px}}
            h1{{color:#3867AE;border-bottom:2px solid #3867AE;padding-bottom:6px}}
            h2{{color:#0F84B5;margin-top:20px}}
            table{{width:100%;border-collapse:collapse;margin:10px 0}}
            th{{background:#3867AE;color:#fff;padding:6px 8px;text-align:right}}
            td{{padding:5px 8px;border-bottom:1px solid #ddd}}
            .badge{{display:inline-block;padding:2px 8px;border-radius:10px;font-size:10px}}
            .yes{{background:#dcfce7;color:#166534}}.no{{background:#f1f5f9;color:#64748b}}
            pre{{background:#f8fafc;border:1px solid #e2e8f0;padding:8px;border-radius:4px;font-size:10px;white-space:pre-wrap;max-height:300px;overflow:auto;direction:ltr;text-align:left}}
        </style></head><body>
        <h1>Mubtkir AI Creator — Template Export</h1>
        <p>Generated: {frappe.utils.now()}</p>
        <p>Total templates: {len(templates)}</p>"""
    ]

    current_type = None
    for t in templates:
        if t.artifact_type != current_type:
            current_type = t.artifact_type
            html_parts.append(f"<h2>{current_type}</h2>")

        deployable = '<span class="badge yes">Deployable</span>' if t.deployable else '<span class="badge no">View only</span>'
        html_parts.append(f"""
            <table>
                <tr><th>Title</th><td>{frappe.utils.escape_html(t.title)}</td><th>Client</th><td>{frappe.utils.escape_html(t.source_client or '')}</td></tr>
                <tr><th>Source Name</th><td>{frappe.utils.escape_html(t.source_name or '')}</td><th>DocType</th><td>{frappe.utils.escape_html(t.target_doctype or '')}</td></tr>
                <tr><th>Version</th><td>{t.version}</td><th>Status</th><td>{deployable}</td></tr>
                <tr><th>Captured</th><td>{t.captured_on or ''}</td><th>Notes</th><td>{frappe.utils.escape_html(t.notes or '')}</td></tr>
            </table>
            <pre>{frappe.utils.escape_html(t.payload or '')[:5000]}</pre>
            <hr>
        """)

    html_parts.append("</body></html>")
    html_content = "\n".join(html_parts)

    fname = f"AI_Templates_{frappe.utils.now().replace(' ', '_').replace(':', '')}.pdf"
    fpath = frappe.get_site_path("private", "files", fname)

    try:
        from frappe.utils.pdf import get_pdf
        pdf_data = get_pdf(html_content)
        with open(fpath, "wb") as f:
            f.write(pdf_data)
    except Exception as e:
        frappe.throw(f"PDF generation failed: {str(e)[:300]}")

    fdoc = frappe.get_doc({
        "doctype": "File",
        "file_name": fname,
        "file_url": f"/private/files/{fname}",
        "is_private": 1,
    })
    fdoc.insert(ignore_permissions=True)
    frappe.db.commit()

    return {"file_url": fdoc.file_url, "file_name": fname, "template_count": len(templates)}


@frappe.whitelist()
def run_export(format="excel", client_site=None, artifact_types=None, target_doctype=None):
    frappe.only_for(["System Manager", "AI Creator User", "AI Creator Supervisor"])
    if isinstance(artifact_types, str):
        try:
            artifact_types = json.loads(artifact_types)
        except ValueError:
            artifact_types = [artifact_types]

    if format == "pdf":
        return export_templates_pdf(client_site, artifact_types, target_doctype)
    return export_templates_excel(client_site, artifact_types, target_doctype)
