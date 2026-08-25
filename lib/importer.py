"""أداة الاستيراد — مطابقة لخطوات Data Import الأصلية في Frappe، بفرق واحد:
التنفيذ يذهب لموقع عميل عبر API بدل قاعدة البيانات المحلية.

- preview(): يقرأ الملف، يبني صفوف التحويل، ويفحص كل صف (حقول إجبارية + حقول
  ربط + صيغة التاريخ) لحفظ قائمة الصفوف الفاشلة سلفًا (invalid_row_numbers) —
  هذا الفحص لا يظهر للمستخدم قبل التنفيذ (تمامًا مثل الأصلي: تفاصيل صحة
  البيانات تظهر فقط بعد Start Import ضمن Import Log)، لكنه يُستخدم لتوفير
  استدعاءات API الضائعة على صفوف معروف فشلها.
- execute(): يتجاوز الصفوف المعروف فشلها تلقائيًا (بلا أي استدعاء API)،
  وينفّذ فقط الباقي — Insert أو Update حسب import_type — ويبني Import Log
  بنفس شكل الأصلي (Row Number / Status / Message / Traceback).
"""

import csv
import datetime
import io
import json
import os
import time

import frappe
from frappe.utils import now_datetime

from mubtkir_ai_creator.lib import llm, tools
from mubtkir_ai_creator.lib.agent import _dump, log_action
from mubtkir_ai_creator.lib.client import FrappeSiteClient

SAMPLE_ROWS = 5
MAX_TOTAL_ROWS = 20000
COMMIT_EVERY = 25
MAX_PREVIEW_ROWS = 100
MAX_LOG_ENTRIES = 500  # حد أقصى لعدد سطور Import Log المحفوظة (حماية حجم القاعدة)

DATE_FORMATS = (
    "%Y-%m-%d", "%Y/%m/%d",
    "%d-%m-%Y", "%d/%m/%Y",
    "%m-%d-%Y", "%m/%d/%Y",
    "%d.%m.%Y", "%Y-%m-%d %H:%M:%S",
)


def _try_parse_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, (datetime.date, datetime.datetime)):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s:
        return None
    for fmt in DATE_FORMATS:
        try:
            return datetime.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


# ---------------- قراءة الملف ----------------

def read_rows(file_url):
    name = frappe.db.get_value("File", {"file_url": file_url}, "name")
    if not name:
        frappe.throw(f"لم يُعثر على الملف: {file_url}")

    doc = frappe.get_doc("File", name)
    ext = os.path.splitext(doc.file_name or "")[1].lower()
    content = doc.get_content()
    raw = content.encode("utf-8") if isinstance(content, str) else content

    if ext in (".xlsx", ".xlsm"):
        headers, rows = _read_xlsx_rows(raw)
    elif ext == ".csv":
        headers, rows = _read_csv_rows(raw)
    else:
        frappe.throw(f"صيغة غير مدعومة للاستيراد: {ext}. المدعوم: xlsx و csv")

    if len(rows) > MAX_TOTAL_ROWS:
        frappe.throw(f"الملف يحتوي {len(rows)} صفًا — الحد الأقصى {MAX_TOTAL_ROWS}")

    return headers, rows


def read_rows_for_import(doc):
    if doc.source_file:
        return read_rows(doc.source_file)
    if doc.google_sheet_url:
        return _read_google_sheet_rows(doc.google_sheet_url)
    frappe.throw("حدد ملف استيراد أو رابط Google Sheet")


def _read_google_sheet_rows(sheet_url):
    import re
    import requests

    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", sheet_url or "")
    if not m:
        frappe.throw("رابط Google Sheet غير صالح — لازم يكون بصيغة docs.google.com/spreadsheets/d/...")
    sheet_id = m.group(1)
    gid_match = re.search(r"[?#&]gid=(\d+)", sheet_url)
    gid = gid_match.group(1) if gid_match else "0"

    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
    try:
        resp = requests.get(export_url, timeout=30)
    except requests.RequestException as e:
        frappe.throw(f"تعذّر الوصول لملف Google Sheet: {e}")

    if resp.status_code != 200 or resp.text.lstrip().lower().startswith("<!doctype html"):
        frappe.throw("تعذّر قراءة Google Sheet — تأكد أن صلاحية المشاركة \"يمكن لأي شخص لديه الرابط العرض\"")

    headers, rows = _read_csv_rows(resp.content)
    if len(rows) > MAX_TOTAL_ROWS:
        frappe.throw(f"الشيت يحتوي {len(rows)} صفًا — الحد الأقصى {MAX_TOTAL_ROWS}")
    return headers, rows


def _read_xlsx_rows(raw):
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    ws = wb.worksheets[0]

    headers, rows = [], []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        values = ["" if c is None else str(c).strip() for c in row]
        if i == 0:
            headers = [h for h in values]
            continue
        if not any(v for v in values):
            continue
        rows.append({headers[j]: values[j] for j in range(min(len(headers), len(values))) if headers[j]})

    wb.close()
    return [h for h in headers if h], rows


def _read_csv_rows(raw):
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("cp1256", errors="replace")

    reader = csv.DictReader(io.StringIO(text))
    headers = [h.strip() for h in (reader.fieldnames or []) if h]
    rows = []
    for r in reader:
        clean = {(k or "").strip(): (str(v).strip() if v is not None else "") for k, v in r.items() if k}
        if any(clean.values()):
            rows.append(clean)
    return headers, rows


# ---------------- بناء الخريطة (الاستدعاء الوحيد للنموذج) ----------------

def analyze(name):
    doc = frappe.get_doc("AI Import", name)
    headers, rows = read_rows_for_import(doc)

    if not headers:
        frappe.throw("لم يُعثر على رؤوس أعمدة في الملف")

    client = FrappeSiteClient(doc.client_site)
    spec = tools.describe_required(client, doc.target_doctype, with_options=True)
    meta = client.get_meta(doc.target_doctype).get("data", {}) or {}

    fields_brief = [
        {
            "fieldname": f.get("fieldname"),
            "label": f.get("label"),
            "fieldtype": f.get("fieldtype"),
            "link_to": f.get("options") if f.get("fieldtype") == "Link" else None,
            "required": bool(f.get("reqd")),
        }
        for f in meta.get("fields", [])
        if f.get("fieldname") and f.get("fieldtype") not in ("Section Break", "Column Break", "Tab Break", "HTML", "Table")
    ][:120]

    sample = rows[:SAMPLE_ROWS]

    id_hint = ""
    if doc.import_type == "Update Existing Records":
        id_hint = '\nهذا استيراد "تحديث سجلات موجودة" — إن وجد عمود يمثل المعرّف/الاسم (ID)، اربطه بحقل "name".'

    prompt = f"""اربط أعمدة ملف الاستيراد بحقول DocType «{doc.target_doctype}» في ERPNext.

أعمدة الملف: {json.dumps(headers, ensure_ascii=False)}

عيّنة من الصفوف: {json.dumps(sample, ensure_ascii=False)}

حقول الـ DocType المتاحة: {json.dumps(fields_brief, ensure_ascii=False)}

الحقول الإجبارية: {json.dumps([f['fieldname'] for f in spec['required_fields']], ensure_ascii=False)}
{id_hint}
أعد JSON فقط بلا أي نص آخر وبلا علامات markdown، بهذا الشكل:
{{
  "mapping": {{"اسم العمود في الملف": "fieldname في ERPNext"}},
  "unmapped_columns": ["أعمدة لم تجد لها مقابلًا"],
  "missing_required": ["حقول إجبارية لا يوجد لها عمود في الملف"],
  "notes": "ملاحظات موجزة بالعربية"
}}"""

    result = llm.chat([{"role": "user", "content": prompt}], system="أنت محلل بيانات. أعد JSON صالحًا فقط.", heavy=True)

    text = (result.get("text") or "").strip()
    text = text.replace("```json", "").replace("```", "").strip()
    try:
        parsed = json.loads(text)
    except ValueError:
        start, end = text.find("{"), text.rfind("}")
        if start == -1 or end == -1:
            frappe.throw(f"تعذّر قراءة خريطة الحقول من النموذج. الرد: {text[:400]}")
        parsed = json.loads(text[start : end + 1])

    doc.db_set("total_rows", len(rows))
    doc.db_set("status", "Mapping Ready")

    mapping = parsed.get("mapping") or {}
    sample_lookup = sample[0] if sample else {}
    doc.set("field_mapping", [])
    for column in headers:
        doc.append(
            "field_mapping",
            {
                "source_column": column,
                "target_fieldname": mapping.get(column, ""),
                "sample_value": str(sample_lookup.get(column, ""))[:140],
            },
        )
    doc.save(ignore_permissions=True)
    frappe.db.commit()

    return {"headers": headers, "total_rows": len(rows), "mapping": mapping, "notes": parsed.get("notes"), "sample_rows": sample, "unmapped": parsed.get("unmapped_columns") or []}


# ---------------- تحويل الصفوف ----------------

def _date_fields_for(client, target_doctype):
    meta = client.get_meta(target_doctype).get("data", {}) or {}
    return {
        f.get("fieldname")
        for f in meta.get("fields", [])
        if f.get("fieldtype") in ("Date", "Datetime") and f.get("fieldname")
    }


def _build_docs(doc, rows, date_fields=None):
    """يبني payload كل صف. في وضع التحديث، يُستخرج المعرّف (name) من الـ payload
    ويُرجَع منفصلاً (row['id']) لاستخدامه في update_doc."""
    mapping = {r.source_column: r.target_fieldname for r in (doc.field_mapping or []) if r.target_fieldname}
    statics = json.loads(doc.static_values or "{}")
    date_fields = date_fields or set()
    is_update = doc.import_type == "Update Existing Records"

    out = []
    for i, row in enumerate(rows, start=2):
        payload = dict(statics)
        date_issues = []
        for column, fieldname in mapping.items():
            if not fieldname:
                continue
            value = row.get(column)
            if value in (None, ""):
                continue
            if fieldname in date_fields:
                parsed_date = _try_parse_date(value)
                if parsed_date:
                    payload[fieldname] = parsed_date
                else:
                    date_issues.append(f"تعذّر فهم التاريخ في {fieldname}: «{value}»")
                    payload[fieldname] = value
            else:
                payload[fieldname] = value

        row_id = payload.pop("name", None) if is_update else None
        out.append({"row_number": i, "data": payload, "id": row_id, "date_issues": date_issues})
    return out


# ---------------- المعاينة ----------------

def preview(name):
    doc = frappe.get_doc("AI Import", name)
    _, rows = read_rows_for_import(doc)
    client = FrappeSiteClient(doc.client_site)
    date_fields = _date_fields_for(client, doc.target_doctype)
    prepared = _build_docs(doc, rows, date_fields)
    is_update = doc.import_type == "Update Existing Records"

    spec = tools.describe_required(client, doc.target_doctype, with_options=True)
    meta = client.get_meta(doc.target_doctype).get("data", {}) or {}
    field_labels = {f.get("fieldname"): f.get("label") or f.get("fieldname") for f in meta.get("fields", [])}

    link_fields = {}
    for f in meta.get("fields", []):
        if f.get("fieldtype") == "Link" and f.get("options"):
            link_fields[f.get("fieldname")] = f.get("options")

    # في وضع التحديث لا تُشترط الحقول الإجبارية — فقط وجود المعرّف
    required = [] if is_update else [f for f in spec["required_fields"] if not f.get("default")]

    unique_values = {}
    for item in prepared:
        for fieldname, target in link_fields.items():
            v = item["data"].get(fieldname)
            if v:
                unique_values.setdefault(fieldname, set()).add(v)

    invalid_links = {}
    for fieldname, values in unique_values.items():
        target = link_fields[fieldname]
        for v in values:
            try:
                found = client.get_list(target, fields=["name"], filters={"name": v}, limit=1).get("data") or []
            except Exception:
                found = []
            if not found:
                invalid_links.setdefault(fieldname, {"doctype": target, "values": []})
                invalid_links[fieldname]["values"].append(v)

    issues, ok_count = [], 0
    invalid_row_reasons = {}
    for item in prepared:
        row_issues = list(item.get("date_issues") or [])
        if is_update and not item.get("id"):
            row_issues.append("لا يوجد معرّف (ID) لهذا الصف — لازم عمود مربوط بحقل name")
        for f in required:
            if not item["data"].get(f["fieldname"]):
                row_issues.append(f"حقل إجباري ناقص: {f.get('label') or f['fieldname']}")
        for fieldname, info in invalid_links.items():
            v = item["data"].get(fieldname)
            if v and v in info["values"]:
                row_issues.append(f"قيمة ربط غير موجودة في {fieldname}: {v}")

        if row_issues:
            issues.append({"row": item["row_number"], "issues": row_issues})
            invalid_row_reasons[str(item["row_number"])] = row_issues
        else:
            ok_count += 1

    # ---- preview_result: أعمدة + صفوف بقيم حقيقية (مربوطة وغير مربوطة) ----
    columns = []
    for r in (doc.field_mapping or []):
        columns.append({
            "source_column": r.source_column,
            "fieldname": r.target_fieldname or None,
            "label": field_labels.get(r.target_fieldname, r.target_fieldname) if r.target_fieldname else None,
            "mapped": bool(r.target_fieldname),
        })

    invalid_rows_by_num = {i["row"]: i["issues"] for i in issues}
    data_rows = []
    for item in prepared[:MAX_PREVIEW_ROWS]:
        row_issues = invalid_rows_by_num.get(item["row_number"])
        # القيم بترتيب أعمدة field_mapping (من الملف الخام، لا من payload، عشان تظهر حتى غير المربوطة)
        raw_row = rows[item["row_number"] - 2]
        cells = [raw_row.get(c["source_column"], "") for c in columns]
        data_rows.append({"row_number": item["row_number"], "values": cells, "ok": not row_issues})

    preview_result = {
        "columns": columns,
        "rows": data_rows,
        "total_rows": len(prepared),
        "max_rows_exceeded": len(prepared) > MAX_PREVIEW_ROWS,
        "max_rows_in_preview": MAX_PREVIEW_ROWS,
    }

    doc.db_set("total_rows", len(prepared))
    doc.db_set("valid_rows", ok_count)
    doc.db_set("invalid_rows", len(issues))
    doc.db_set("preview_result", json.dumps(preview_result, ensure_ascii=False))
    doc.db_set("invalid_row_numbers", json.dumps(invalid_row_reasons, ensure_ascii=False))
    doc.db_set("status", "Pending Approval")
    frappe.db.commit()

    return {"valid": ok_count, "invalid": len(issues)}


@frappe.whitelist()
def get_template(client_site, target_doctype):
    """يبني ملف CSV بعناوين الحقول الإجبارية (والاسم لو تحديث سجلات) لتحميله كقالب."""
    frappe.only_for(["System Manager", "AI Creator User", "AI Creator Supervisor"])
    client = FrappeSiteClient(client_site)
    spec = tools.describe_required(client, target_doctype, with_options=False)

    headers = [f.get("label") or f.get("fieldname") for f in spec["required_fields"]]
    if not headers:
        meta = client.get_meta(target_doctype).get("data", {}) or {}
        headers = [
            f.get("label") or f.get("fieldname")
            for f in meta.get("fields", [])
            if f.get("fieldtype") not in ("Section Break", "Column Break", "Tab Break", "HTML", "Table")
        ][:15]

    buf = io.StringIO()
    csv.writer(buf).writerow(headers)

    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": f"{target_doctype}-template.csv",
        "content": buf.getvalue(),
        "is_private": 1,
    })
    file_doc.insert(ignore_permissions=True)
    frappe.db.commit()
    return {"file_url": file_doc.file_url}


# ---------------- التنفيذ ----------------

def enqueue_execute(name):
    doc = frappe.get_doc("AI Import", name)
    if doc.status != "Approved":
        frappe.throw("لا يمكن التنفيذ قبل الاعتماد")

    doc.db_set("status", "Queued")
    doc.db_set("processed_rows", 0)
    frappe.db.commit()

    frappe.enqueue(
        "mubtkir_ai_creator.lib.importer.execute",
        queue="long",
        timeout=7200,
        name=name,
        enqueue_after_commit=True,
    )
    return {"status": "Queued"}


def _parse_remote_error(err_text):
    """يحاول استخراج عنوان/تفصيل مقروء من نص خطأ الـ API — يُبقي النص الخام كـ traceback."""
    title, detail = "Failed", err_text[:300]
    try:
        start = err_text.find("{")
        if start != -1:
            payload = json.loads(err_text[start:])
            msgs = payload.get("_server_messages")
            if msgs:
                parsed_msgs = json.loads(msgs)
                first = json.loads(parsed_msgs[0]) if parsed_msgs else {}
                detail = first.get("message") or detail
            exc_type = payload.get("exc_type") or ""
            if "DuplicateEntry" in exc_type or "already exists" in detail:
                title = "Duplicate Name"
            elif "ValidationError" in exc_type:
                title = "Validation Error"
            elif "LinkValidationError" in exc_type or "MandatoryError" in exc_type:
                title = "Missing / Invalid Value"
            elif "PermissionError" in exc_type:
                title = "Permission Error"
    except Exception:
        pass
    return title, detail


def execute(name):
    started = time.time()
    doc = frappe.get_doc("AI Import", name)
    _, rows = read_rows_for_import(doc)
    client = FrappeSiteClient(doc.client_site)
    date_fields = _date_fields_for(client, doc.target_doctype)
    prepared = _build_docs(doc, rows, date_fields)
    is_update = doc.import_type == "Update Existing Records"

    known_invalid = json.loads(doc.invalid_row_numbers or "{}")
    skip_invalid = bool(doc.skip_invalid_rows)

    is_submittable = False
    if doc.submit_after_import:
        meta = client.get_meta(doc.target_doctype).get("data", {}) or {}
        is_submittable = bool(meta.get("is_submittable"))

    doc.db_set("status", "Executing")
    frappe.db.commit()

    success = failed = skipped_known = 0
    log = []

    for idx, item in enumerate(prepared, start=1):
        row_key = str(item["row_number"])
        if row_key in known_invalid:
            failed += 1
            skipped_known += 1
            if len(log) < MAX_LOG_ENTRIES:
                log.append({
                    "row": item["row_number"], "status": "Failure",
                    "title": "Skipped (Invalid at Preview)",
                    "detail": " | ".join(known_invalid[row_key]),
                    "traceback": None,
                })
        else:
            try:
                if is_update:
                    out = client.update_doc(doc.target_doctype, item["id"], item["data"])
                    created_name = item["id"]
                else:
                    out = client.create_doc(doc.target_doctype, item["data"])
                    created_name = (out or {}).get("data", {}).get("name")

                if is_submittable and created_name:
                    try:
                        full_doc = client.get_doc(doc.target_doctype, created_name).get("data") or {}
                        client.call_method("frappe.client.submit", {"doc": json.dumps(full_doc)})
                    except Exception as sub_err:
                        # فشل الاعتماد لا يُلغي نجاح الإنشاء — يُسجَّل كملاحظة فقط
                        if len(log) < MAX_LOG_ENTRIES:
                            log.append({
                                "row": item["row_number"], "status": "Failure",
                                "title": "Created but Not Submitted",
                                "detail": str(sub_err)[:300],
                                "traceback": str(sub_err)[:2000],
                            })
                success += 1
            except Exception as e:
                failed += 1
                title, detail = _parse_remote_error(str(e))
                if not skip_invalid:
                    doc.db_set("status", "Failed")
                    doc.db_set("error_message", f"توقف عند الصف {item['row_number']}: {detail}")
                if len(log) < MAX_LOG_ENTRIES:
                    log.append({
                        "row": item["row_number"], "status": "Failure",
                        "title": title, "detail": detail, "traceback": str(e)[:2000],
                    })
                if not skip_invalid:
                    break

        if idx % COMMIT_EVERY == 0:
            doc.db_set("processed_rows", idx)
            doc.db_set("success_count", success)
            doc.db_set("failed_count", failed)
            frappe.db.commit()

    doc.db_set("processed_rows", len(prepared))
    doc.db_set("success_count", success)
    doc.db_set("failed_count", failed)
    doc.db_set("import_log", json.dumps(log, ensure_ascii=False))

    if doc.status != "Failed":
        doc.db_set("status", "Completed" if failed == 0 else "Partially Failed")

    log_action(
        client_site=doc.client_site,
        site_url=client.site_url,
        tool_name="bulk_import",
        risk_level="high",
        tool_input=_dump({"import": doc.name, "doctype": doc.target_doctype, "import_type": doc.import_type, "rows": len(prepared)}),
        tool_output=_dump({"success": success, "failed": failed, "skipped_known_invalid": skipped_known}),
        is_success=1 if failed == 0 else 0,
        duration_ms=int((time.time() - started) * 1000),
        error_message=None if failed == 0 else f"فشل {failed} صفًا ({skipped_known} منها متجاوَز من المعاينة) — راجع Import Log",
    )

    frappe.db.commit()
    return {"status": doc.status, "success": success, "failed": failed, "skipped_known_invalid": skipped_known}
