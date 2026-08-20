"""Template Service — generate/download import templates for remote DocTypes."""

import io
import json

import frappe

from mubtkir_ai_creator.lib.remote_import.metadata import get_doctype_meta


def generate_template(client_site, doctype, with_data=False):
    """Generate an Excel import template based on remote DocType metadata.

    Returns a dict with:
      - headers: list of column header strings
      - child_headers: dict of {table_fieldname: [headers]}
      - sample_row: optional sample data row
    """
    meta = get_doctype_meta(client_site, doctype)

    headers = ["name"]  # Always include name column for Update operations
    field_map = {}

    for f in meta["fields"]:
        fn = f["fieldname"]
        ft = f["fieldtype"]
        label = f.get("label", fn)
        reqd = " *" if f.get("reqd") else ""

        if ft in ("Attach", "Attach Image"):
            headers.append(fn)
            field_map[fn] = f
        elif ft not in ("Table", "Table MultiSelect"):
            headers.append(fn)
            field_map[fn] = f

    child_headers = {}
    for table_fn, table_info in meta.get("child_tables", {}).items():
        child_dt = table_info["doctype"]
        ch_headers = ["name"]
        for cf in table_info["fields"]:
            ch_headers.append(cf["fieldname"])
        child_headers[table_fn] = {
            "doctype": child_dt,
            "headers": ch_headers,
            "fields": table_info["fields"],
        }

    return {
        "doctype": doctype,
        "headers": headers,
        "fields": meta["fields"],
        "child_headers": child_headers,
        "is_submittable": meta.get("is_submittable", 0),
    }


def download_template_excel(client_site, doctype):
    """Generate and return an Excel file as bytes for the import template."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        frappe.throw("openpyxl is required — pip install openpyxl")

    tmpl = generate_template(client_site, doctype)

    wb = Workbook()
    ws = wb.active
    ws.title = doctype[:31]  # Excel tab name limit

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    reqd_fill = PatternFill(start_color="C75050", end_color="C75050", fill_type="solid")

    # Write parent headers
    for col, h in enumerate(tmpl["headers"], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        # Check if field is required
        field_info = next((f for f in tmpl["fields"] if f["fieldname"] == h), None)
        if field_info and field_info.get("reqd"):
            cell.fill = reqd_fill
        else:
            cell.fill = hdr_fill
        ws.column_dimensions[cell.column_letter].width = max(len(h) + 4, 15)

    # Add child table sheets
    for table_fn, ch_info in tmpl.get("child_headers", {}).items():
        ch_ws = wb.create_sheet(title=ch_info["doctype"][:31])
        for col, h in enumerate(ch_info["headers"], 1):
            cell = ch_ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            field_info = next((f for f in ch_info["fields"] if f["fieldname"] == h), None)
            if field_info and field_info.get("reqd"):
                cell.fill = reqd_fill
            else:
                cell.fill = hdr_fill
            ch_ws.column_dimensions[cell.column_letter].width = max(len(h) + 4, 15)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
