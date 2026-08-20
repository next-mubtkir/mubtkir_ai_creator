"""Preview Engine — parse uploaded files, validate, and preview data before import."""

import csv
import io
import json
import os

import frappe


def parse_google_sheet(url):
    """Parse a public Google Sheet by converting to CSV export URL."""
    import requests

    # Convert Google Sheets URL to CSV export
    # Format: https://docs.google.com/spreadsheets/d/SHEET_ID/...
    sheet_id = None
    if "/spreadsheets/d/" in url:
        parts = url.split("/spreadsheets/d/")[1].split("/")
        sheet_id = parts[0]

    if not sheet_id:
        frappe.throw("Invalid Google Sheet URL — must contain /spreadsheets/d/")

    csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"

    try:
        resp = requests.get(csv_url, timeout=30)
        resp.raise_for_status()
    except Exception as e:
        frappe.throw(f"Failed to load Google Sheet: {e}. Make sure the sheet is Public")

    return parse_file(file_content=resp.content, file_name=f"google_sheet_{sheet_id}.csv")


def parse_file(file_url=None, file_content=None, file_name=None):
    """Parse an uploaded Excel or CSV file and return headers + rows.

    Returns:
        {
            "headers": [col1, col2, ...],
            "rows": [[val1, val2, ...], ...],
            "total_rows": int,
            "file_name": str,
            "file_type": "xlsx" | "csv",
        }
    """
    if file_url:
        file_path = frappe.get_site_path("public" if "/public/" in file_url else "", file_url.lstrip("/"))
        if not os.path.exists(file_path):
            file_path = frappe.get_site_path(file_url.lstrip("/"))
        if not os.path.exists(file_path):
            # Try as private file
            file_doc = frappe.get_doc("File", {"file_url": file_url})
            file_path = file_doc.get_full_path()

        file_name = file_name or os.path.basename(file_path)
        with open(file_path, "rb") as f:
            file_content = f.read()

    if not file_content:
        frappe.throw("No file specified for preview")

    file_name = file_name or "unknown"
    ext = file_name.rsplit(".", 1)[-1].lower() if "." in file_name else ""

    if ext in ("xlsx", "xls"):
        return _parse_excel(file_content, file_name)
    elif ext == "csv":
        return _parse_csv(file_content, file_name)
    else:
        # Try CSV first, then Excel
        try:
            return _parse_csv(file_content, file_name)
        except Exception:
            return _parse_excel(file_content, file_name)


def _parse_excel(content, file_name):
    """Parse Excel file."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        frappe.throw("openpyxl is required — pip install openpyxl")

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows = []
    headers = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        row_vals = [_clean_cell(c) for c in row]
        if i == 0:
            headers = row_vals
        else:
            # Skip completely empty rows
            if any(v for v in row_vals):
                rows.append(row_vals)

    wb.close()

    return {
        "headers": headers,
        "rows": rows,
        "total_rows": len(rows),
        "file_name": file_name,
        "file_type": "xlsx",
    }


def _parse_csv(content, file_name):
    """Parse CSV file."""
    if isinstance(content, bytes):
        # Try UTF-8 first, then windows-1256 (common for Arabic)
        for enc in ("utf-8-sig", "utf-8", "cp1256", "latin-1"):
            try:
                text = content.decode(enc)
                break
            except (UnicodeDecodeError, ValueError):
                continue
        else:
            frappe.throw("Could not detect file encoding")
    else:
        text = content

    reader = csv.reader(io.StringIO(text))
    headers = []
    rows = []
    for i, row in enumerate(reader):
        row_vals = [c.strip() for c in row]
        if i == 0:
            headers = row_vals
        else:
            if any(v for v in row_vals):
                rows.append(row_vals)

    return {
        "headers": headers,
        "rows": rows,
        "total_rows": len(rows),
        "file_name": file_name,
        "file_type": "csv",
    }


def preview_data(file_url, limit=20):
    """Parse file and return a preview (first N rows)."""
    result = parse_file(file_url=file_url)
    return {
        "headers": result["headers"],
        "rows": result["rows"][:limit],
        "total_rows": result["total_rows"],
        "file_name": result["file_name"],
        "file_type": result["file_type"],
        "preview_rows": min(limit, result["total_rows"]),
    }


def validate_data(file_url, mapping, client_site, doctype):
    """Validate mapped data against remote DocType constraints.

    Returns a list of warnings/errors per row.
    """
    from mubtkir_ai_creator.lib.remote_import.metadata import get_doctype_meta

    result = parse_file(file_url=file_url)
    meta = get_doctype_meta(client_site, doctype)

    # Build field lookup
    field_lookup = {}
    for f in meta["fields"]:
        field_lookup[f["fieldname"]] = f

    if isinstance(mapping, str):
        mapping = json.loads(mapping)

    warnings = []
    for row_idx, row in enumerate(result["rows"], start=2):  # 2 = Excel row (header is 1)
        row_warnings = []
        for col_idx, header in enumerate(result["headers"]):
            target_field = mapping.get(header)
            if not target_field or col_idx >= len(row):
                continue

            value = row[col_idx]
            field_info = field_lookup.get(target_field)
            if not field_info:
                continue

            # Check required
            if field_info.get("reqd") and not value:
                row_warnings.append(f"Required field '{field_info.get('label', target_field)}' is empty")

            # Check data types
            if value and field_info["fieldtype"] in ("Int", "Float", "Currency", "Percent"):
                try:
                    float(str(value).replace(",", ""))
                except (ValueError, TypeError):
                    row_warnings.append(f"القيمة '{value}' ليست رقمية للحقل '{target_field}'")

        if row_warnings:
            warnings.append({"row": row_idx, "warnings": row_warnings})

    return {
        "total_rows": result["total_rows"],
        "warnings_count": len(warnings),
        "warnings": warnings[:100],  # Cap at 100
    }


def _clean_cell(value):
    """Clean a cell value for consistency."""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        # Remove .0 from integers stored as float
        if isinstance(value, float) and value == int(value):
            return str(int(value))
        return str(value)
    return str(value).strip()
